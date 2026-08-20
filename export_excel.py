"""
Libro de Excel de asistencia (v2, con descansos).
Marca Elite VA Consulting: Dorado #A36F06 (color exacto del logo oficial), tipografía Poppins.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DORADO = "A36F06"
IVORY = "FAFAFA"
NEAR_BLACK = "0C0C0C"
GREY = "6B6B6B"
ROJO = "D62612"

# Poppins es la tipografía de marca. Si no está instalada en el equipo del
# lector, Excel sustituye automáticamente. Cambia a "Arial" si prefieres.
BRAND_FONT = "Poppins"
DATA_FONT = "Poppins"

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_FECHA = "DD/MM/YYYY"
FMT_HORA = "DD/MM/YYYY HH:MM"
FMT_HORAS = "0.00"
FMT_ENTERO = "#,##0"


def _encabezado(ws, row: int, headers: list[str]):
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.font = Font(name=BRAND_FONT, bold=True, size=10, color=IVORY)
        cell.fill = PatternFill("solid", fgColor=DORADO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _anchos(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_workbook(resumen: list[dict], sesiones: list[dict],
                   inicio: date, fin: date, actor: str, alcance: str) -> BytesIO:
    wb = Workbook()

    # ================================================================
    # Hoja 1 — Resumen por empleado
    # ================================================================
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Control de Asistencia"
    ws["B2"].font = Font(name=BRAND_FONT, bold=True, size=22, color=DORADO)
    ws["B3"] = "Entrada · Descanso · Salida"
    ws["B3"].font = Font(name=BRAND_FONT, size=11, color=DORADO)

    meta = [
        ("Periodo", f"{inicio:%d/%m/%Y} — {fin:%d/%m/%Y}"),
        ("Alcance", alcance),
        ("Generado para", actor),
        ("Generado el", f"{date.today():%d/%m/%Y}"),
        ("Días con registro", len(resumen)),
    ]
    for i, (label, value) in enumerate(meta, start=5):
        ws.cell(row=i, column=2, value=label).font = Font(
            name=BRAND_FONT, bold=True, size=10, color=GREY)
        ws.cell(row=i, column=3, value=value).font = Font(
            name=DATA_FONT, size=10, color=NEAR_BLACK)

    ws["B12"] = "Totales por empleado"
    ws["B12"].font = Font(name=BRAND_FONT, bold=True, size=13, color=DORADO)

    personas = sorted({
        (r["nombre"], r["email"], r.get("departamento") or "—") for r in resumen
    })
    _encabezado(ws, 13, ["Empleado", "Correo", "Departamento", "Días",
                         "Horas presencia", "Horas descanso", "Horas trabajadas",
                         "Horas extra", "Horas faltantes", "Exceso descanso (min)",
                         "Promedio trabajado/día"])

    for idx, (nombre, email, dept) in enumerate(personas):
        r = 14 + idx
        ws.cell(row=r, column=1, value=nombre)
        ws.cell(row=r, column=2, value=email)
        ws.cell(row=r, column=3, value=dept)
        clave = f"$B{r}"
        # En la hoja Detalle: C=Correo, H=Presencia, I=Descanso, J=Trabajadas,
        #                     M=Exceso descanso, N=Extra, O=Faltantes
        ws.cell(row=r, column=4,  value=f"=COUNTIF(Detalle!$C:$C,{clave})")
        ws.cell(row=r, column=5,  value=f"=SUMIF(Detalle!$C:$C,{clave},Detalle!$H:$H)")
        ws.cell(row=r, column=6,  value=f"=SUMIF(Detalle!$C:$C,{clave},Detalle!$I:$I)")
        ws.cell(row=r, column=7,  value=f"=SUMIF(Detalle!$C:$C,{clave},Detalle!$J:$J)")
        ws.cell(row=r, column=8,  value=f"=SUMIF(Detalle!$C:$C,{clave},Detalle!$N:$N)")
        ws.cell(row=r, column=9,  value=f"=SUMIF(Detalle!$C:$C,{clave},Detalle!$O:$O)")
        ws.cell(row=r, column=10, value=f"=SUMIF(Detalle!$C:$C,{clave},Detalle!$M:$M)")
        ws.cell(row=r, column=11, value=f"=IFERROR($G{r}/$D{r},0)")

        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=DATA_FONT, size=10, color=NEAR_BLACK)
            cell.border = BORDER
            if c in (5, 6, 7, 8, 9, 11):
                cell.number_format = FMT_HORAS
            elif c in (4, 10):
                cell.number_format = FMT_ENTERO

    if personas:
        t = 14 + len(personas)
        for c in range(1, 12):
            cell = ws.cell(row=t, column=c)
            cell.fill = PatternFill("solid", fgColor=DORADO)
            cell.border = BORDER
            cell.font = Font(name=BRAND_FONT, bold=True, size=10, color=IVORY)
            if c == 1:
                cell.value = "TOTAL"
            elif c in (4, 5, 6, 7, 8, 9, 10):
                col = get_column_letter(c)
                cell.value = f"=SUM({col}14:{col}{t - 1})"
                cell.number_format = FMT_ENTERO if c in (4, 10) else FMT_HORAS

    _anchos(ws, {1: 26, 2: 30, 3: 16, 4: 8, 5: 15, 6: 15, 7: 16,
                 8: 13, 9: 15, 10: 19, 11: 20})

    # ================================================================
    # Hoja 2 — Detalle diario
    # ================================================================
    ws2 = wb.create_sheet("Detalle")
    cols = ["Fecha", "Empleado", "Correo", "Departamento", "Primera entrada",
            "Última salida", "Sesiones", "Horas presencia", "Horas descanso",
            "Horas trabajadas", "Descanso permitido (min)", "Trabajo esperado",
            "Exceso descanso (min)", "Horas extra", "Horas faltantes"]
    _encabezado(ws2, 1, cols)

    for i, row in enumerate(resumen, start=2):
        valores = [
            row["fecha"], row["nombre"], row["email"], row.get("departamento"),
            row.get("primera_entrada"), row.get("ultima_salida"), row["sesiones"],
            row["horas_presencia"], row["horas_descanso"], row["horas_trabajadas"],
            row["descanso_permitido_min"], row["trabajo_esperado"],
            row["descanso_excedido_min"], row["horas_extra"], row["horas_faltantes"],
        ]
        for c, v in enumerate(valores, start=1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font = Font(name=DATA_FONT, size=10, color=NEAR_BLACK)
            cell.border = BORDER
            if c == 1:
                cell.number_format = FMT_FECHA
            elif c in (5, 6):
                cell.number_format = FMT_HORA
            elif c in (8, 9, 10, 12, 14, 15):
                cell.number_format = FMT_HORAS
            elif c in (7, 11, 13):
                cell.number_format = FMT_ENTERO
            # Resaltar el exceso de descanso
            if c == 13 and v is not None and float(v) > 0:
                cell.font = Font(name=DATA_FONT, size=10, bold=True, color=ROJO)

    if resumen:
        ref = f"A1:{get_column_letter(len(cols))}{len(resumen) + 1}"
        tabla = Table(displayName="DetalleDiario", ref=ref)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight8", showRowStripes=True, showColumnStripes=False)
        ws2.add_table(tabla)
    _anchos(ws2, {1: 12, 2: 24, 3: 30, 4: 15, 5: 18, 6: 18, 7: 10, 8: 15,
                  9: 15, 10: 16, 11: 20, 12: 16, 13: 19, 14: 13, 15: 15})

    # ================================================================
    # Hoja 3 — Sesiones (cada jornada individual)
    # ================================================================
    ws3 = wb.create_sheet("Sesiones")
    cols3 = ["Fecha", "Empleado", "Departamento", "Entrada", "Salida",
             "Horas presencia", "Horas descanso", "Horas trabajadas", "Descansos"]
    _encabezado(ws3, 1, cols3)

    for i, row in enumerate(sesiones, start=2):
        valores = [row["fecha"], row["nombre"], row.get("departamento"),
                   row["entrada"], row["salida"], row["horas_presencia"],
                   row["horas_descanso"], row["horas_trabajadas"], row["descansos"]]
        for c, v in enumerate(valores, start=1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.font = Font(name=DATA_FONT, size=10, color=NEAR_BLACK)
            cell.border = BORDER
            if c == 1:
                cell.number_format = FMT_FECHA
            elif c in (4, 5):
                cell.number_format = FMT_HORA
            elif c in (6, 7, 8):
                cell.number_format = FMT_HORAS
            elif c == 9:
                cell.number_format = FMT_ENTERO

    if sesiones:
        ref3 = f"A1:{get_column_letter(len(cols3))}{len(sesiones) + 1}"
        tabla3 = Table(displayName="Sesiones", ref=ref3)
        tabla3.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight8", showRowStripes=True, showColumnStripes=False)
        ws3.add_table(tabla3)
    _anchos(ws3, {1: 12, 2: 24, 3: 16, 4: 18, 5: 18, 6: 15, 7: 15, 8: 16, 9: 11})

    # ================================================================
    # Hoja 4 — Notas
    # ================================================================
    ws4 = wb.create_sheet("Notas")
    ws4.sheet_view.showGridLines = False
    ws4["B2"] = "Cómo leer este reporte"
    ws4["B2"].font = Font(name=BRAND_FONT, bold=True, size=14, color=DORADO)

    notas = [
        ("Cómo se calculan las horas",
         "Presencia = tiempo entre entrada y salida. Descanso = suma de los periodos "
         "marcados como descanso. Trabajadas = presencia menos descanso. "
         "Solo las horas trabajadas cuentan para extras y faltantes."),
        ("Jornada y descanso son por persona",
         "Cada empleado tiene su propia jornada y su propio descanso permitido, "
         "configurables por el administrador. La columna 'Trabajo esperado' muestra la "
         "meta de trabajo efectivo que aplica a esa persona."),
        ("Jornada que incluye el descanso",
         "Por defecto, una jornada de 8 h con 1 h de descanso significa 8 h de presencia "
         "y 7 h de trabajo efectivo. Si un empleado está configurado al revés, su "
         "'Trabajo esperado' será de 8 h y su presencia esperada de 9 h."),
        ("Exceso de descanso",
         "Minutos de descanso por encima de lo permitido, resaltados en rojo. Es "
         "informativo: no se descuenta automáticamente de las horas trabajadas."),
        ("Sesiones",
         "Un día puede tener varias sesiones si la persona marcó salida y volvió a entrar. "
         "El detalle diario suma todas las sesiones de la fecha."),
        ("Jornadas sin cerrar",
         "Si alguien olvida marcar salida, esa jornada NO aparece aquí hasta que se "
         "corrija. Revisa las alertas del panel antes de cerrar el periodo."),
        ("Cambios de horario",
         "La jornada y el descanso se leen de la configuración ACTUAL del empleado. "
         "Si cambias el horario de alguien, los días anteriores se recalculan con el "
         "valor nuevo. Exporta y archiva el reporte antes de hacer cambios retroactivos."),
        ("Correcciones y auditoría",
         "Las marcaciones manuales quedan con registro de quién las hizo y por qué; no se "
         "borra el registro original. Cada exportación también queda registrada."),
    ]
    fila = 4
    for titulo, cuerpo in notas:
        ws4.cell(row=fila, column=2, value=titulo).font = Font(
            name=BRAND_FONT, bold=True, size=10, color=DORADO)
        c = ws4.cell(row=fila, column=3, value=cuerpo)
        c.font = Font(name=DATA_FONT, size=10, color=NEAR_BLACK)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws4.row_dimensions[fila].height = 44
        fila += 2
    _anchos(ws4, {2: 30, 3: 95})

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
